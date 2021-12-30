import time
import datetime as dt
import openpyxl.worksheet.copier
import pandas as pd
from openpyxl import load_workbook
from openpyxl import workbook, Workbook
from openpyxl import utils
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
import pathlib
from openpyxl.styles.protection import Protection


EXPORT_FOLDER_PATH = "./Export"
FORMAT_SHEET = "Format.xlsx"
DATA_FILE = "1.xlsx"
COLORS_LIST = ["00FFFF00","00FF00FF","0000FFFF",
"00FF0000","0000FF00","000000FF","00FFFF00","00FF00FF",
"0000FFFF","00800000","00008000","00000080","00808000",
"00800080","00008080","00C0C0C0","00808080","009999FF"]
now = dt.datetime.now()
curent_time = now.strftime("%Y_%m_%d")

def locate_column(target_worksheet,header):
    for column, cellObj in enumerate(target_worksheet[11]):
        if cellObj.value == header: return column

def comments(source_file,target_file):
    """find all comments in ws and assign them to new ws"""
    wb = load_workbook(source_file)
    ws =wb.active
    comments_dic = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                key = row[2].value
                comments_dic.setdefault(key,[])
                comments_dic[key].append([ws.cell(column=cell.column,row=11).value , cell.comment])
    wb2= load_workbook(target_file)
    sheets = wb2.sheetnames
    for sheet in sheets[3:]:
        ws2 = wb2[sheet]
        for index,row in enumerate(ws2.iter_rows()):
            if row[2].value in comments_dic:
                worker_id = row[2].value
                for comment in comments_dic[worker_id]:
                    column = locate_column(ws2,comment[0])
                    ws2.cell(column=column+1,row=index+1).comment = comment[1]
    wb2.save(target_file)

def write_on_every_line(worksheet,text, column):
    """adding formulas for every row:"""
    for row, cellObj in enumerate(list(worksheet.columns)[column]):
        if row >= 11:
            n = text.format(row+1)
            # n = '=(E%d-F%d)/C%d' % (row+1, row+1, row+1)
            cellObj.value = n

def create_formulas(source_file):
    """create formulas for insertion"""
    wb= load_workbook(source_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=13,max_row=13):
        formulas_dic = {}
        for column, cellObj in enumerate(row):
            if cellObj.data_type == "f" : formulas_dic [ws[11][column].value] =cellObj.value
        for key, value in formulas_dic.items():
            formulas_dic[key] = value.replace("13","{0}")
        return formulas_dic

def insert_formulas(target_file):
    """adds formulas to file"""
    wb = load_workbook(target_file)
    sheets = wb.sheetnames
    for sheet in sheets[3:]:
        formulas_dic = create_formulas()
        for key,value in formulas_dic.items():
            ws = wb[sheet]
            column = locate_column(ws,key)
            write_on_every_line(ws,value,column)
    wb.save(target_file)

def lock_sheet(source_file,target_file):
    """Locks all sheets execp allowed tables """
    open_columns = []
    wb = load_workbook(source_file)
    ws = wb.active
    for column,cell in enumerate(ws[12]):
        if cell.fill.start_color.index == 0:
            open_columns.append(column)
    wb2 = load_workbook(target_file)
    sheets = wb2.sheetnames
    for sheet in sheets:
        ws2 = wb2[sheet]
        ws2.protection.sheet = True
        for index in open_columns:
            column = openpyxl.utils.get_column_letter(index+1)
            for cell in ws[column]:
                if cell.row >= 12:
                    cell.protection = Protection(locked=False)
    wb2.save(target_file)

def create_df():
    """creates df with headers and formulas as string"""
    wb_data = load_workbook("1.xlsx")
    ws_data = wb_data.active
    df2 = pd.read_excel("1.xlsx")
    headers = df2.iloc[9].values
    df = pd.DataFrame(ws_data.values)
    # the drop is because the panda recognize empty column but th openpyxl not/ so i drop i manualy
    df = df.drop(columns=103)
    df.columns = headers
    df = df[11:]   #my main data starts on row 12
    df = df.reset_index(drop=True)
    return df

def create_beauty(amount_of_rows,amount_of_columns,worksheet):
    """makes a same cell format. to a data only area"""
    for row in range(12,amount_of_rows+13):
        for cell in range(1,amount_of_columns):
            source_cell = worksheet.cell(column=cell, row=12)
            target_cell = worksheet.cell(column=cell, row=row)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)
    worksheet.delete_rows(12)

def create_file_director(name,list_of_managers):
    """Creates directory and a file according to a format for director"""
    file_folder = f"Export/{name}"
    pathlib.Path(file_folder).mkdir(parents=True, exist_ok=True)
    wb = load_workbook(FORMATE)
    ws = wb["format"]
    for manager in list_of_managers:
        ws2 = wb.copy_worksheet(ws)
        ws2.title = manager
        ws2.sheet_properties.tabColor = COLORS_LIST[list_of_managers.index(manager)]
        ws2.sheet_view.showGridLines = False
    del wb["format"]
    wb.save(f"{file_folder}/{name}.xlsx")
    return file_folder

def create_file_manager(folder,name):
    """Creates directory and a file according to a format for manager"""
    file_folder = f"{folder}/{name}"
    pathlib.Path(file_folder).mkdir(parents=True, exist_ok=True)
    wb = load_workbook(FORMATE)
    ws = wb["format"]
    ws2 = wb.copy_worksheet(ws)
    ws2.title = name
    ws2.sheet_properties.tabColor = COLORS_LIST[1]
    ws2.sheet_view.showGridLines = False
    del wb["format"]
    wb.save(f"{file_folder}/{name}.xlsx")
    return file_folder

start_time = time.perf_counter()
df = create_df()
director_list = list(df["Director"].unique())
for director in director_list:
    df_director = df.loc[df["Director"] == director]
    group_managers = list(df_director["Group Manager"].unique())
    director_folder = create_file_director(director,group_managers)
    wb = load_workbook(f"{director_folder}/{director}.xlsx")
    for manager in group_managers:
        ws_director = wb[manager]
        manager_folder = create_file_manager(director_folder,manager)
        wb_manager = load_workbook(f"{manager_folder}/{manager}.xlsx")
        ws_manager = wb_manager[manager]
        df_manager=df_director.loc[df_director["Group Manager"]==manager]
        for r in dataframe_to_rows(df_manager, index=False, header=False):
            ws_director.append(r)
            ws_manager.append(r)
        create_beauty(df_manager["Group Manager"].size, len(df_director.columns), ws_director)
        create_beauty(df_manager["Group Manager"].size, len(df_director.columns),ws_manager)
        wb_manager.save(f"{manager_folder}/{manager}.xlsx")
    wb.save(f"{director_folder}/{director}.xlsx")
    director_time = time.perf_counter()
    print(f"Time to director {director} --- %s seconds ---"% (director_time - start_time))
    insert_formulas(f"{director_folder}/{director}.xlsx")
    comments(DATA_FILE, f"{director_folder}/{director}.xlsx")
    lock_sheet(f"{director_folder}/{director}.xlsx")
end_time = time.perf_counter()
print("TOTAL TIME --- %s seconds ---" % (end_time - start_time))
